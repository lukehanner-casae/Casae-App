#!/usr/bin/env python3
"""
One-off import of real property/lodger data from the "Casae Living - Rental
Properties.xlsx" workbook (Lodgers sheet) into Supabase.

    python3 scripts/import-lodgers.py [--xlsx PATH] [--apply]

Without --apply it is a dry run: the generated SQL is executed inside a DO
block that raises at the end, so Postgres validates and rolls back everything,
and the plan is printed. With --apply the same block runs to completion.

Needs `openpyxl` and the Supabase CLI linked to the project. The workbook is
never copied anywhere; the generated SQL (which contains names/emails) is
written to a scratch path and deleted afterwards.

Rules (see HANDOFF.md, Session L):
  * Properties are matched by normalised street address, rooms by the
    "Bedroom N" number. Anything that doesn't match is reported, not created.
  * Existing lodgers are matched by normalised first + last name; "TBC"
    placeholders left in rooms the sheet fills are deleted.
  * Closed rows, or Open rows whose Lodging End has passed, become `former`.
    Open rows with a future Lodging Start become `pending`. Open rows with a
    future Lodging End become `current` and get a vacate notice logged.
  * Bond received date = lodging start (capped at today) — the sheet's
    "Security Deposits Held" totals confirm every Open bond is held.
  * occupancy_history is rebuilt per room from the lodging intervals
    (source = 'backfill'), replacing the seed-time snapshot rows.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import subprocess
import sys
import tempfile
from collections import defaultdict
from zoneinfo import ZoneInfo

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

PERTH = ZoneInfo("Australia/Perth")
TODAY = dt.datetime.now(PERTH).date()
DEFAULT_XLSX = os.path.expanduser("~/Downloads/Casae Living - Rental Properties.xlsx")


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def norm_address(s: str) -> str:
    s = s.split(",")[0].lower()
    s = re.sub(r"\b(\d+)[a-z]\b", r"\1", s)          # 332D -> 332
    s = re.sub(r"\bstreet\b", "st", s)
    s = re.sub(r"\broad\b", "rd", s)
    s = re.sub(r"\bavenue\b", "ave", s)
    return re.sub(r"\s+", " ", s).strip()


def norm_name(s: str | None) -> str:
    return re.sub(r"[^a-z0-9 ]", "", (s or "").lower().replace("’", "")).strip()


def split_people(name: str) -> list[str]:
    parts = re.split(r"\s+(?:&|and)\s+", name.strip(), flags=re.I)
    return [p.strip() for p in parts if p.strip()]


def first_last(person: str) -> tuple[str, str | None]:
    tokens = person.split()
    return tokens[0], (" ".join(tokens[1:]) or None)


def as_date(v) -> dt.date | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    return dt.date.fromisoformat(str(v)[:10])


def lit(v) -> str:
    """SQL literal."""
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, dt.date):
        return f"'{v.isoformat()}'"
    return "'" + str(v).replace("'", "''") + "'"


def db_query(sql: str) -> list[dict]:
    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False) as f:
        f.write(sql)
        path = f.name
    try:
        out = subprocess.run(
            ["supabase", "db", "query", "--linked", "--file", path],
            capture_output=True, text=True,
        )
    finally:
        os.unlink(path)
    if out.returncode != 0:
        raise RuntimeError(out.stdout + out.stderr)
    text = out.stdout[out.stdout.index("{"):] if "{" in out.stdout else "{}"
    return json.loads(text).get("rows", [])


def db_exec(sql: str) -> str:
    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False) as f:
        f.write(sql)
        path = f.name
    try:
        out = subprocess.run(
            ["supabase", "db", "query", "--linked", "--file", path],
            capture_output=True, text=True,
        )
    finally:
        os.unlink(path)
    return out.stdout + out.stderr


# ---------------------------------------------------------------------------
# load
# ---------------------------------------------------------------------------

def load_sheet(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Lodgers"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for n, row in enumerate(rows[1:], start=2):
        get = lambda k: row[idx[k]] if k in idx and idx[k] < len(row) else None
        if not get("Property"):
            continue
        out.append({
            "row": n,
            "property": str(get("Property")).strip(),
            "room": str(get("Room") or "").strip(),
            "name": (str(get("Lodger Name")).strip() if get("Lodger Name") else None),
            "email": (str(get("Email")).strip() if get("Email") else None),
            "n_lodgers": get("# Lodgers"),
            "start": as_date(get("Lodging Start")),
            "bond": get("Bond Amount"),
            "end": as_date(get("Lodging End")),
            "bond_return": as_date(get("Bond Return Date")),
            "status": (str(get("Status")).strip() if get("Status") else None),
            "notes": (str(get("Notes")).strip() if get("Notes") else None),
            "agreement": (str(get("Lodging Agreement")).strip() if get("Lodging Agreement") else None),
        })
    # "Total Deposits Held" block at the bottom of the Properties sheet.
    deposits = {}
    wp = wb["Properties"]
    capture = False
    for row in wp.iter_rows(values_only=True):
        vals = list(row)
        if any(v == "Total Deposits Held" for v in vals):
            capture = True
            continue
        if capture and vals[1] and isinstance(vals[4], (int, float)):
            deposits[str(vals[1]).strip()] = float(vals[4])
    return out, deposits


def load_db():
    props = db_query("select id, display_name, address, weekly_head_lease, head_lease_start from properties where status = 'active' order by display_name;")
    rooms = db_query("select id, property_id, room_name, status, weekly_rent from rooms;")
    lodgers = db_query("select id, first_name, last_name, room_id, status, email, bond_amount from lodgers;")
    return props, rooms, lodgers


# ---------------------------------------------------------------------------
# plan
# ---------------------------------------------------------------------------

def build_plan(rows, deposits, props, rooms, lodgers):
    flags: list[str] = []
    prop_by_addr = {norm_address(p["address"] or ""): p for p in props}
    room_by = {(r["property_id"], r["room_name"]): r for r in rooms}
    rooms_by_prop = defaultdict(list)
    for r in rooms:
        rooms_by_prop[r["property_id"]].append(r)

    matched_props: dict[str, dict] = {}      # sheet name -> db property
    unmatched_props: dict[str, list[int]] = defaultdict(list)
    plan = []                                # per-row actions
    used_lodger_ids: dict[str, int] = {}     # db lodger id -> sheet row

    for row in rows:
        p = prop_by_addr.get(norm_address(row["property"]))
        if not p:
            unmatched_props[row["property"]].append(row["row"])
            continue
        matched_props[row["property"]] = p
        m = re.search(r"bedroom\s*(\d+)", row["room"], re.I)
        room = room_by.get((p["id"], f"Room {m.group(1)}")) if m else None
        if not room:
            flags.append(f"row {row['row']}: room '{row['room']}' not found in {p['display_name']} — skipped")
            continue
        if not row["name"]:
            plan.append({"row": row, "property": p, "room": room, "kind": "empty"})
            continue

        people = split_people(row["name"])
        first, last = first_last(people[0])
        partner = people[1] if len(people) > 1 else None
        is_couple = bool(partner) or (row["n_lodgers"] or 0) >= 2
        emails = [e.strip() for e in re.split(r"[\n,;]+", row["email"] or "") if e.strip()]

        # status
        s = (row["status"] or "").lower()
        if s == "closed" or (row["end"] and row["end"] <= TODAY):
            status = "former"
            if s == "open":
                flags.append(f"row {row['row']}: {row['name']} is Open but Lodging End {row['end']} has passed — imported as former")
        elif row["start"] and row["start"] > TODAY:
            status = "pending"
        else:
            status = "current"
        notice_date = row["end"] if (status == "current" and row["end"] and row["end"] > TODAY) else None
        if s not in ("open", "closed"):
            flags.append(f"row {row['row']}: unexpected Status '{row['status']}' — treated as {status}")
        if row["bond"] in (None, 0):
            flags.append(f"row {row['row']}: {row['name']} has no bond amount ({row['bond']!r})")

        # match existing lodger (by name; the DB has no emails yet)
        cand = []
        for l in lodgers:
            if norm_name(l["first_name"]) != norm_name(first):
                continue
            ln = norm_name(l["last_name"])
            if ln and last and not norm_name(last).endswith(ln) and not ln.endswith(norm_name(last).split()[-1]):
                continue
            cand.append(l)
        existing = None
        if len(cand) == 1:
            existing = cand[0]
        elif len(cand) > 1:
            # prefer same room, then same property
            same_room = [c for c in cand if c["room_id"] == room["id"]]
            same_prop = [c for c in cand if c["room_id"] and room_by_id(rooms, c["room_id"])["property_id"] == p["id"]]
            existing = (same_room or same_prop or [None])[0]
            if existing is None:
                flags.append(f"row {row['row']}: {row['name']} matches several existing lodgers — inserting new")
        if existing and existing["id"] in used_lodger_ids:
            flags.append(f"row {row['row']}: {row['name']} — existing lodger already claimed by row {used_lodger_ids[existing['id']]}; inserting new")
            existing = None
        if existing:
            used_lodger_ids[existing["id"]] = row["row"]

        notes = []
        stamp = TODAY.strftime("%-d %b %Y")
        if row["notes"]:
            notes.append(f"[{stamp} — import] {row['notes']}")
        if len(emails) > 1:
            notes.append(f"[{stamp} — import] Partner email: {', '.join(emails[1:])}")
        if row["agreement"]:
            notes.append(f"[{stamp} — import] Agreement file: {row['agreement']}")

        bond = float(row["bond"]) if row["bond"] not in (None, "") else None
        bond_received = min(row["start"], TODAY) if (bond and row["start"]) else None
        plan.append({
            "row": row, "property": p, "room": room, "kind": "update" if existing else "insert",
            "existing": existing,
            "fields": {
                "first_name": first, "last_name": last, "email": emails[0] if emails else None,
                "room_id": room["id"], "move_in_date": row["start"], "expected_move_out": row["end"],
                "bond_amount": bond, "bond_received_date": bond_received,
                "bond_returned_date": row["bond_return"] if status == "former" else None,
                "lodging_agreement_signed": bool(row["agreement"]),
                "is_couple": is_couple, "partner_name": partner, "status": status,
                "notes": "\n".join(notes) or None,
            },
            "notice_date": notice_date,
        })

    # existing lodgers in matched properties that no sheet row claimed
    matched_prop_ids = {p["id"] for p in matched_props.values()}
    leftovers = []
    for l in lodgers:
        if l["id"] in used_lodger_ids or not l["room_id"]:
            continue
        r = room_by_id(rooms, l["room_id"])
        if r["property_id"] not in matched_prop_ids:
            continue
        name = f"{l['first_name'] or ''} {l['last_name'] or ''}".strip()
        if norm_name(l["first_name"]) == "tbc":
            leftovers.append({"lodger": l, "action": "delete", "why": "seed placeholder"})
        else:
            leftovers.append({"lodger": l, "action": "former", "why": f"not in sheet ({name})"})
            flags.append(f"existing lodger '{name}' ({room_label(props, rooms, r)}) is not in the sheet — set to former, not deleted")

    return plan, leftovers, matched_props, unmatched_props, flags


def room_by_id(rooms, rid):
    return next(r for r in rooms if r["id"] == rid)


def room_label(props, rooms, r):
    p = next(p for p in props if p["id"] == r["property_id"])
    return f"{p['display_name']} / {r['room_name']}"


# ---------------------------------------------------------------------------
# occupancy timeline per room (from lodging intervals, up to today)
# ---------------------------------------------------------------------------

def room_timeline(intervals: list[tuple[dt.date, dt.date | None]]) -> list[tuple[dt.date, str]]:
    """Returns [(date, 'occupied'|'vacant')...] transitions, chronological."""
    points = set()
    for s, e in intervals:
        if s and s <= TODAY:
            points.add(s)
        if e and e < TODAY:
            points.add(e + dt.timedelta(days=1))
    events = []
    occupied = False
    for d in sorted(points):
        now_occ = any(s and s <= d and (e is None or e >= d) for s, e in intervals)
        if now_occ != occupied:
            events.append((d, "occupied" if now_occ else "vacant"))
            occupied = now_occ
    return events


# ---------------------------------------------------------------------------
# SQL
# ---------------------------------------------------------------------------

def perth_ts(d: dt.date) -> str:
    return f"({lit(d)}::timestamp at time zone 'Australia/Perth')"


def build_sql(plan, leftovers, rooms, props, dry_run: bool) -> tuple[str, dict]:
    stmts = []
    stmts.append("alter table rooms disable trigger rooms_log_status_change;")

    for lo in leftovers:
        if lo["action"] == "delete":
            stmts.append(f"delete from lodgers where id = {lit(lo['lodger']['id'])};")
        else:
            stmts.append(f"update lodgers set status = 'former' where id = {lit(lo['lodger']['id'])};")

    notices = []
    for a in plan:
        if a["kind"] == "empty":
            continue
        f = a["fields"]
        cols = ", ".join(f.keys())
        if a["kind"] == "update":
            sets = ", ".join(f"{k} = {lit(v)}" for k, v in f.items())
            stmts.append(f"update lodgers set {sets} where id = {lit(a['existing']['id'])};")
            if a["notice_date"]:
                notices.append((a, f"(select id from lodgers where id = {lit(a['existing']['id'])})"))
        else:
            vals = ", ".join(lit(v) for v in f.values())
            stmts.append(f"insert into lodgers ({cols}) values ({vals});")
            if a["notice_date"]:
                notices.append((a, f"(select id from lodgers where email = {lit(f['email'])} and room_id = {lit(f['room_id'])} order by move_in_date desc limit 1)"))

    # room state + history, per matched room
    by_room: dict[str, list] = defaultdict(list)
    for a in plan:
        by_room[a["room"]["id"]].append(a)
    room_state = {}
    history_rows = 0
    for rid, actions in by_room.items():
        residents = [a for a in actions if a["kind"] != "empty" and a["fields"]["status"] in ("current", "pending")]
        formers = [a for a in actions if a["kind"] != "empty" and a["fields"]["status"] == "former"]
        intervals = [(a["fields"]["move_in_date"], a["fields"]["expected_move_out"]) for a in actions if a["kind"] != "empty" and a["fields"]["move_in_date"]]
        events = room_timeline(intervals)
        if residents:
            db_room = room_by_id(rooms, rid)
            status = "notice_given" if db_room["status"] == "notice_given" else "occupied"
            since = None
        else:
            last_end = max((a["fields"]["expected_move_out"] for a in formers if a["fields"]["expected_move_out"]), default=None)
            status, since = "vacant", last_end
        r = room_by_id(rooms, rid)
        prop = next(p for p in props if p["id"] == r["property_id"])
        lease_start = as_date(prop.get("head_lease_start"))
        if status == "vacant" and since is None and lease_start:
            since = lease_start  # never occupied: vacant since the head lease began
        room_state[rid] = (status, since, events)
        stmts.append(
            f"update rooms set status = {lit(status)}, vacant_since = {perth_ts(since) if since else 'null'} where id = {lit(rid)};"
        )
        stmts.append(f"delete from occupancy_history where room_id = {lit(rid)} and source = 'backfill';")
        prev = None
        for d, st in events:
            stmts.append(
                "insert into occupancy_history (room_id, property_id, status, previous_status, changed_at, source) values "
                f"({lit(rid)}, {lit(r['property_id'])}, {lit(st)}, {lit(prev)}, {perth_ts(d)}, 'backfill');"
            )
            prev = st
            history_rows += 1
        if not events:
            # No past occupancy: vacant from the head-lease start (or now if unknown).
            stmts.append(
                "insert into occupancy_history (room_id, property_id, status, previous_status, changed_at, source) values "
                f"({lit(rid)}, {lit(r['property_id'])}, 'vacant', null, {perth_ts(lease_start) if lease_start else 'now()'}, 'backfill');"
            )
            history_rows += 1
            if status == "occupied":  # held by a pending (future) move-in only
                stmts.append(
                    "insert into occupancy_history (room_id, property_id, status, previous_status, changed_at, source) values "
                    f"({lit(rid)}, {lit(r['property_id'])}, 'occupied', 'vacant', now(), 'backfill');"
                )
                history_rows += 1

    stmts.append("alter table rooms enable trigger rooms_log_status_change;")
    stmts.append(
        "update lodgers l set status = 'notice_given' where l.status = 'current' "
        "and exists (select 1 from vacate_notices vn where vn.lodger_id = l.id and vn.status = 'active');"
    )
    for a, lodger_sub in notices:
        stmts.append(
            f"if not exists (select 1 from vacate_notices where lodger_id = {lodger_sub} and status = 'active') then "
            f"perform log_vacate_notice({lit(a['property']['id'])}, {lit(a['room']['id'])}, {lodger_sub}, "
            f"{lit(a['notice_date'])}, 'Imported from spreadsheet: Lodging End {a['notice_date']}'); end if;"
        )
    if dry_run:
        stmts.append("raise exception 'DRY RUN — rolled back';")

    body = "\n  ".join(stmts)
    sql = f"do $import$\nbegin\n  {body}\nend $import$;\n"
    return sql, {"room_state": room_state, "history_rows": history_rows, "notices": len(notices)}


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    rows, deposits = load_sheet(args.xlsx)
    props, rooms, lodgers = load_db()
    plan, leftovers, matched, unmatched, flags = build_plan(rows, deposits, props, rooms, lodgers)
    sql, info = build_sql(plan, leftovers, rooms, props, dry_run=not args.apply)

    print(f"Sheet rows: {len(rows)} · today (Perth): {TODAY}\n")
    print("MATCHED PROPERTIES")
    for sheet_name, p in matched.items():
        print(f"  {sheet_name:28} -> {p['display_name']}")
    print("\nUNMATCHED PROPERTIES (skipped, not created)")
    for name, rws in unmatched.items():
        print(f"  {name:28} rows {rws}")
    print("\nLODGER ACTIONS")
    for a in plan:
        r = a["row"]
        label = f"{a['property']['display_name']} / {a['room']['room_name']}"
        if a["kind"] == "empty":
            print(f"  row {r['row']:>2} {label:24} (no lodger listed)")
            continue
        f = a["fields"]
        who = f"{f['first_name']} {f['last_name'] or ''}".strip() + (f" & {f['partner_name']}" if f["partner_name"] else "")
        ex = f" (updates '{a['existing']['first_name']} {a['existing']['last_name'] or ''}'.strip())" if a["kind"] == "update" else ""
        ex = ex.replace(".strip()", "")
        extra = f" + vacate notice {a['notice_date']}" if a["notice_date"] else ""
        print(f"  row {r['row']:>2} {label:24} {a['kind']:6} {who:45} {f['status']:8} bond {f['bond_amount'] or 0:>6.0f}{ex}{extra}")
    print("\nEXISTING LODGERS NOT IN SHEET")
    for lo in leftovers:
        l = lo["lodger"]
        print(f"  {lo['action']:6} {l['first_name']} {l['last_name'] or ''} — {lo['why']}")
    print("\nROOM STATE AFTER IMPORT")
    for rid, (status, since, events) in info["room_state"].items():
        r = room_by_id(rooms, rid)
        ev = ", ".join(f"{d}:{s}" for d, s in events) or "no history"
        print(f"  {room_label(props, rooms, r):24} {status:8} {('since ' + since.isoformat()) if since else '':16} history: {ev}")

    # bond cross-check against the sheet's "Security Deposits Held"
    print("\nBOND CROSS-CHECK (held after import vs sheet 'Security Deposits Held')")
    held = defaultdict(float)
    for a in plan:
        if a["kind"] != "empty" and a["fields"]["status"] != "former" and a["fields"]["bond_amount"]:
            held[a["row"]["property"]] += a["fields"]["bond_amount"]
    for name in matched:
        sheet = deposits.get(name)
        ok = "OK" if sheet is not None and abs(sheet - held[name]) < 0.01 else "MISMATCH"
        print(f"  {name:28} import {held[name]:>7.0f}  sheet {sheet if sheet is not None else '?':>7}  {ok}")

    print("\nFLAGS")
    for fl in flags or ["(none)"]:
        print(f"  - {fl}")
    print(f"\nSQL: {sql.count(chr(10))} lines, {info['history_rows']} history rows, {info['notices']} vacate notice(s)")

    print("\n" + ("APPLYING…" if args.apply else "DRY RUN (validating against the database, then rolling back)…"))
    out = db_exec(sql)
    if args.apply:
        if "error" in out.lower() and "rows" not in out:
            print(out)
            sys.exit(1)
        print("Applied.")
    else:
        if "DRY RUN" in out:
            print("Dry run OK — SQL validated and rolled back.")
        else:
            print(out)
            sys.exit(1)


if __name__ == "__main__":
    main()
